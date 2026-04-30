#!/usr/bin/env python
"""Q3 LEGAL dev pre-review XLSX generator.

Propósito
=========
Genera un XLSX con las 23 strings marcadas LEGAL DRAFT en `i18n/eu.po`,
reducido a solo el source ES, para que el dev haga una pasada de
detección de problemas obvios (datos incorrectos, typos, coherencia)
ANTES de mandar el paquete real a asesoría legal externa.

Diferencia vs `gen_q1_xlsx.py`:
  - 1 sola pestaña «LEGAL DRAFT (ES)» (no 7 bloques temáticos).
  - Columnas reducidas: ID | ES (source) | Notas dev | Contexto.
    No se incluye EU DRAFT ni EU FINAL — esta fase es solo ES; la
    validación lingüística EU se separa en Q1.
  - Filtro: solo entries con marker `LEGAL DRAFT - REVIEW NEEDED` en el
    translator comment.

Reusa la lógica de cálculo de row heights del script Q1 para evitar
text overlap.

Requisitos: polib + openpyxl (mismo venv que Q1).
"""
import hashlib
import sys
import textwrap

import polib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PO_PATH = '/opt/odoo/v14/workspace/website_avanzosc_demo/i18n/eu.po'
OUT_PATH = ('/opt/odoo/v14/workspace/website_avanzosc_demo/'
            'docs/q3-legal-validation/dev-prereview-2026-04-30/'
            'avanzosc-q3-legal-prereview-2026-04-30.xlsx')

LINE_HEIGHT_PT = 15
PADDING_PT = 4
PROPORTIONAL_FACTOR = 1.2

# Map ref keys → human label (subset of the gen_q1 map, only legal pages)
LABEL_MAP = {
    'page_aviso_legal': '/aviso-legal',
    'page_politica_privacidad': '/politica-privacidad',
    'page_politica_cookies': '/politica-cookies',
}


def make_id(msgid):
    h = hashlib.sha1(msgid.encode('utf-8')).hexdigest()[:6].upper()
    return f'Q3-{h}'


def compute_row_height(cell_texts, col_widths_chars):
    max_lines = 1
    for text, col_w in zip(cell_texts, col_widths_chars):
        text = str(text or '')
        if not text:
            continue
        effective_w = max(1, int(col_w * PROPORTIONAL_FACTOR))
        segments = text.split('\n')
        total = 0
        for seg in segments:
            wrapped = textwrap.wrap(seg, width=effective_w,
                                    break_long_words=True, replace_whitespace=False)
            total += len(wrapped) or 1
        max_lines = max(max_lines, total)
    return max_lines * LINE_HEIGHT_PT + PADDING_PT


def context_text(entry):
    refs = [f for f, _ in (entry.occurrences or [])]
    keys = []
    for r in refs:
        if 'website_avanzosc_demo.' in r:
            keys.append(r.split('website_avanzosc_demo.', 1)[1])
    labels = []
    seen = set()
    for k in keys:
        lbl = LABEL_MAP.get(k, k)
        if lbl not in seen:
            labels.append(lbl)
            seen.add(lbl)
    ctx = ' · '.join(labels)
    tc = (entry.tcomment or '').strip()
    if 'LEGAL DRAFT - REVIEW NEEDED' in tc:
        parts = tc.split('—', 1)
        if len(parts) > 1:
            note = parts[1].strip()
            if note:
                ctx = f'{ctx} — {note}' if ctx else note
    return ctx or '(sin contexto)'


def main():
    po = polib.pofile(PO_PATH)
    legal_entries = [
        e for e in po
        if 'LEGAL DRAFT - REVIEW NEEDED' in (e.tcomment or '')
    ]
    print(f'Found {len(legal_entries)} LEGAL DRAFT entries')

    wb = Workbook()
    wb.remove(wb.active)

    write_instructions(wb, len(legal_entries))
    write_legal_sheet(wb, legal_entries)

    wb.save(OUT_PATH)
    print(f'OK → {OUT_PATH}')


def write_instructions(wb, total):
    ws = wb.create_sheet('INSTRUCCIONES', 0)
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 100

    title_font = Font(bold=True, size=14)
    h2_font = Font(bold=True, size=12, color='1B2A41')
    fill_red = PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid')
    fill_yellow = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')

    rows = [
        ('Avanzosc · Q3 — Pre-revisión legal (dev)', title_font, None),
        ('', None, None),
        (f'Generado: 2026-04-30 · Total strings LEGAL DRAFT: {total}', None, None),
        ('', None, None),
        ('1. ¿Qué es esto?', h2_font, None),
        ('Las 3 páginas legales del sitio (Aviso Legal, Política de Privacidad, Política de Cookies) están redactadas como BORRADOR LEGAL DRAFT generado por el dev. Antes de mandar el paquete a asesoría legal externa, hacemos una pasada interna del dev para detectar problemas obvios (datos incorrectos, typos, incoherencias, secciones que sobran o faltan, links rotos).', None, None),
        ('', None, None),
        ('2. ¿Qué tienes que hacer?', h2_font, None),
        ('  · Lee la columna ES (source) — el texto tal y como aparece en la página renderizada.', None, None),
        ('  · Compara contra los PDFs de la carpeta pdfs/ (vista pública desktop) si necesitas ver el contexto visual completo.', None, None),
        ('  · Anota en la columna Notas dev cualquier problema que detectes:', None, None),
        ('      · Dato sensible incorrecto (CIF mal, dirección, email, teléfono, jurisdicción).', None, None),
        ('      · Typo o error gramatical.', None, None),
        ('      · Sección obvia que falta (ej. derecho oposición RGPD, registro mercantil).', None, None),
        ('      · Sección que sobra (algo que no aplica al servicio real).', None, None),
        ('      · Incoherencia entre las 3 páginas (ej. CIF distinto en aviso vs privacidad).', None, None),
        ('      · Links rotos a otra página o a recurso externo.', None, None),
        ('  · La columna Contexto te dice DÓNDE aparece la string (URL/sección).', None, None),
        ('', None, None),
        ('3. Datos sensibles a verificar (lista separada)', h2_font, None),
        ('Ver `datos-sensibles-extraidos.md` en la misma carpeta — tabla con un checkbox por dato. Marcar manualmente los confirmados; abrir incidencia para los problemáticos.', None, None),
        ('', None, None),
        ('4. Esto NO sustituye revisión legal profesional', h2_font, fill_red),
        ('La asesoría legal externa de Avanzosc DEBE revisar los textos antes de switchover (gate Q3). Esta pre-revisión solo busca problemas detectables sin competencia legal: datos mal, typos, coherencia interna, links rotos. NO valida que los textos cumplan RGPD/LSSI/etc.', fill_red, None),
        ('', None, None),
        ('5. Cómo devolver el archivo', h2_font, None),
        ('  · Guardar el XLSX con las notas en la columna Notas dev.', None, None),
        ('  · Comunicar al equipo: lista de issues detectados que requieran código (templates) vs lista de issues solo para asesoría (textos).', None, None),
        ('  · El dev aplica los fixes de templates antes de mandar el paquete real a asesoría.', None, None),
    ]
    instructions_col_widths = (4, 100)
    for i, (b, font, fill) in enumerate(rows, start=1):
        ws.cell(row=i, column=2, value=b)
        if font:
            ws.cell(row=i, column=2).font = font
        if fill:
            ws.cell(row=i, column=2).fill = fill
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        ws.row_dimensions[i].height = compute_row_height(('', b), instructions_col_widths)


def write_legal_sheet(wb, entries):
    ws = wb.create_sheet('LEGAL DRAFT (ES)')
    ws.sheet_properties.tabColor = 'FFB3B3'

    bold = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2A41', end_color='1B2A41', fill_type='solid')
    fill_notas = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fill_legal = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title
    ws.cell(row=1, column=1, value='LEGAL DRAFT (ES) — Pre-revisión dev').font = Font(bold=True, size=14, color='1B2A41')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    # Description
    ws.cell(row=2, column=1, value='23 strings marcadas LEGAL DRAFT en i18n/eu.po. Solo source ES (la validación EU se separa en Q1).').font = Font(italic=True, size=10, color='555555')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
    # Banner
    ws.cell(row=3, column=1, value='⚠ Esto NO sustituye revisión legal profesional. Anotar problemas de datos/typos/coherencia; la asesoría externa valida cumplimiento normativo.').font = Font(bold=True, size=11, color='AA0000')
    ws.cell(row=3, column=1).fill = fill_legal
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=4)

    header_row = 4
    headers = ['ID', 'ES (source)', 'Notas dev', 'Contexto']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

    widths = [12, 70, 50, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    merged_w = sum(widths)
    ws.row_dimensions[1].height = compute_row_height(('LEGAL DRAFT (ES) — Pre-revisión dev',) + ('',) * 3,
                                                       (merged_w,) + (0,) * 3)
    ws.row_dimensions[2].height = compute_row_height(('23 strings marcadas LEGAL DRAFT en i18n/eu.po. Solo source ES (la validación EU se separa en Q1).',) + ('',) * 3,
                                                       (merged_w,) + (0,) * 3)
    ws.row_dimensions[3].height = compute_row_height(
        ('⚠ Esto NO sustituye revisión legal profesional. Anotar problemas de datos/typos/coherencia; la asesoría externa valida cumplimiento normativo.',) + ('',) * 3,
        (merged_w,) + (0,) * 3)
    ws.row_dimensions[header_row].height = LINE_HEIGHT_PT + PADDING_PT

    entries_sorted = sorted(entries, key=lambda e: (context_text(e), e.msgid.lower()))

    for row_idx, entry in enumerate(entries_sorted, start=header_row + 1):
        cell_values = (
            make_id(entry.msgid),
            entry.msgid,
            '',
            context_text(entry),
        )
        for col, v in enumerate(cell_values, start=1):
            ws.cell(row=row_idx, column=col, value=v)
        ws.cell(row=row_idx, column=3).fill = fill_notas
        for col in range(1, 5):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border
        ws.row_dimensions[row_idx].height = compute_row_height(cell_values, tuple(widths))

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


if __name__ == '__main__':
    main()
