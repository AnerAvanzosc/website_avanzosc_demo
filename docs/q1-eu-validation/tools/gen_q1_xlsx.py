#!/usr/bin/env python
"""Q1 EU validation XLSX generator.

Propósito
=========
Generar un paquete entregable XLSX para el equipo lingüístico de Avanzosc
con todas las strings DRAFT del fichero `i18n/eu.po`, agrupadas por bloques
temáticos. El revisor escribe en columnas EU FINAL y Notas; el dev parsea
de vuelta para mergear al .po (procedimiento detallado en
`docs/q1-validation-runbook.md` §5.2).

Requisitos técnicos
===================
- Python 3.10+ (venv Odoo `/opt/odoo/v14/venv`).
- `polib` (parser .po) — ya disponible en el venv Odoo.
- `openpyxl` 3.1+ (escritor XLSX) — instalable vía:
    /opt/odoo/v14/venv/bin/pip install openpyxl
  Esta dependencia es del SCRIPT, NO del módulo Odoo. El módulo
  `website_avanzosc_demo` no importa openpyxl en runtime; el sysadmin
  no debe preocuparse en producción.

Cómo invocar
============
    /opt/odoo/v14/venv/bin/python docs/q1-eu-validation/tools/gen_q1_xlsx.py

Asunciones sobre `i18n/eu.po`
=============================
- Las entries pendientes de revisión llevan el flag `#, fuzzy` Y un
  comentario translator (`#`) con uno de los markers canónicos:
    `# DRAFT - REVIEW NEEDED — <descripción>`        (gate Q1)
    `# LEGAL DRAFT - REVIEW NEEDED — <descripción>`  (gate Q3)
- Las entries VALIDADAS no llevan flag fuzzy ni marker canónico; el
  filtro es por substring exacto «DRAFT - REVIEW NEEDED» o
  «LEGAL DRAFT - REVIEW NEEDED» para evitar falsos positivos sobre
  entries con comentarios tipo «VALIDADO, NO DRAFT».
- Las occurrences (`#:`) usan el formato Odoo:
    `model_terms:ir.ui.view,arch_db:website_avanzosc_demo.<template_id>`
    `model:website.menu,name:website_avanzosc_demo.<menu_id>`
  El sufijo tras `website_avanzosc_demo.` se usa para asignar bloque
  temático (ver `BLOCKS` abajo).

Salida
======
Archivo: `docs/q1-eu-validation/avanzosc-eu-validation-2026-04-30.xlsx`
(ruta hardcoded en `OUT_PATH` — ajustar fecha al regenerar).

Estructura del XLSX:
- Sheet 0: INSTRUCCIONES — readme inline para el revisor.
- Sheets 1-7: bloques temáticos. Columnas:
    ID  |  ES (source)  |  EU DRAFT (current)  |  EU FINAL  |  Notas  |  Contexto
- ID estable: `Q1-{6-hex-uppercase}` derivado de
  `sha1(msgid).hexdigest()[:6].upper()`. Independiente del orden del .po.

Re-ejecutar al regenerar el paquete
===================================
Si el .po cambia (nuevas strings, correcciones a markers) y se necesita
un paquete nuevo: ajustar `OUT_PATH` con la fecha actual, ejecutar.
El XLSX previo se mantiene como histórico.

Decisiones de bloque temático: ver `BLOCKS` abajo y la sección §5 del
runbook para el racional. Si añades nuevos templates al módulo, asegúrate
de que su prefijo aparece en alguna entrada de `BLOCKS` o caen al
fallback `B8_OTROS` (que el revisor verá como pestaña separada).
"""
import hashlib
import sys
import textwrap

import polib
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

PO_PATH = '/opt/odoo/v14/workspace/website_avanzosc_demo/i18n/eu.po'
OUT_PATH = '/opt/odoo/v14/workspace/website_avanzosc_demo/docs/q1-eu-validation/avanzosc-eu-validation-2026-04-30.xlsx'

# ---------------------------------------------------------------------------
# Row-height calculation constants (added 2026-04-30 fix to prevent text
# overlap; openpyxl leaves row_dimensions[N].height = None by default and
# wrapText=True does NOT auto-resize the row in OOXML — Excel/LibreOffice
# render at default ~15pt and wrapped text overflows behind the next row).
#
# LINE_HEIGHT_PT: 11pt font + line-height ≈ 1.36 ≈ 15pt per visible line.
# PADDING_PT: 4pt of breathing room (top + bottom combined).
# PROPORTIONAL_FACTOR: monospace counting under-estimates how many chars
# fit per line in proportional fonts (Inter/Calibri default in Excel).
# Empirically 1.2 produces a tight fit without significant whitespace.
# ---------------------------------------------------------------------------
LINE_HEIGHT_PT = 15
PADDING_PT = 4
PROPORTIONAL_FACTOR = 1.2

# Block definitions: (block_id, title, sheet_color, template_prefixes_or_keys, description)
# Order matters: first match wins on multi-reference entries.
BLOCKS = [
    ('B7_LEGAL', 'Legal — ESPERAR Q3', 'FFB3B3', (),
     'Bloque LEGAL DRAFT. Esperar revisión Q3 por asesoría legal antes de tocar. Solo informativo.'),
    ('B5_CONTACTO', 'Contacto', 'CCE5FF', (
        'page_contacto', 'page_contacto_gracias', 's_avanzosc_cta_contacto',
    ), 'Página /contacto + /contacto/gracias + CTA contacto transversal.'),
    ('B4_CONOCENOS', 'Conócenos / Equipo', 'D9EAD3', (
        'page_conocenos', 's_avanzosc_equipo',
    ), 'Página Conócenos + snippet equipo. La página /trabaja-con-nosotros'
       ' se eliminó post-v1 (sesión 2026-04-30) y sus 21 strings dejaron'
       ' de estar en el .po.'),
    ('B3_SOLUCIONES', 'Soluciones (4 sectoriales)', 'FFF2CC', (
        'page_industrial', 'page_distribucion', 'page_servicios', 'page_academias',
        's_avanzosc_sector_specifics', 's_avanzosc_caso_exito',
    ), '4 páginas sectoriales + snippet sector_specifics + caso éxito archetype 1.'),
    ('B6_TRANSVERSALES', 'Transversales', 'EAD1DC', (
        's_avanzosc_cta_kit_consulting',
    ), 'Snippets transversales (CTA Kit Consulting) que aparecen en múltiples páginas.'),
    ('B1_HERO_HOME', 'Hero & Home', 'D0E0E3', (
        's_avanzosc_hero', 's_avanzosc_pilares', 's_avanzosc_sectores',
        's_avanzosc_contador', 's_avanzosc_timeline',
    ), 'Snippets visibles en la home: hero principal, pilares, sectores grid, contador, timeline.'),
    ('B2_NAV_FOOTER', 'Navegación & Footer', 'D9D2E9', (
        'menu_', 'footer_avanzosc', 'header_acceso_clientes',
        'header_mobile_buttons', 'layout_mobile_overlay', 'footer_copyright_avanzosc',
    ), 'Menú principal (top-level + dropdown) + header buttons + footer columns.'),
]
# Catch-all for entries that don't match any block above
FALLBACK_BLOCK = ('B8_OTROS', 'Otros (sin bloque explícito)', 'EFEFEF', (),
                  'Strings cuya referencia no coincide con ningún bloque temático declarado. Revisar caso por caso.')


def compute_row_height(cell_texts, col_widths_chars):
    """Calculate row height in points to fit all wrapped text without overlap.

    Args:
      cell_texts: tuple of cell text strings, one per column in this row.
      col_widths_chars: tuple of column widths in chars, same length as cell_texts.

    Algorithm:
      For each non-empty cell, count required visual lines as:
        1 line per explicit '\\n' in text (multi-line cells like INSTRUCCIONES)
        + lines required for wrapping each segment at effective_width chars
      Take the maximum across all columns. Multiply by LINE_HEIGHT_PT and
      add PADDING_PT.

    Returns: float (points). Minimum LINE_HEIGHT_PT + PADDING_PT for
    empty rows (covers spacers).
    """
    max_lines = 1
    for text, col_w in zip(cell_texts, col_widths_chars):
        text = str(text or '')
        if not text:
            continue
        effective_w = max(1, int(col_w * PROPORTIONAL_FACTOR))
        # Split on explicit newlines first, then wrap each segment.
        segments = text.split('\n')
        total = 0
        for seg in segments:
            wrapped = textwrap.wrap(seg, width=effective_w,
                                    break_long_words=True, replace_whitespace=False)
            total += len(wrapped) or 1
        max_lines = max(max_lines, total)
    return max_lines * LINE_HEIGHT_PT + PADDING_PT


def make_id(msgid):
    """Stable Q1-XXXXXX id derived from SHA1 of msgid."""
    h = hashlib.sha1(msgid.encode('utf-8')).hexdigest()[:6].upper()
    return f'Q1-{h}'


def reference_keys(entry):
    """Extract template/menu keys from #: occurrences."""
    keys = []
    for filename, _ in (entry.occurrences or []):
        # filename comes like 'model_terms:ir.ui.view,arch_db:website_avanzosc_demo.page_industrial'
        # or 'model:website.menu,name:website_avanzosc_demo.menu_academias'
        if 'website_avanzosc_demo.' in filename:
            keys.append(filename.split('website_avanzosc_demo.', 1)[1])
        else:
            keys.append(filename)
    return keys


def is_legal(entry):
    """LEGAL DRAFT marker takes precedence over generic DRAFT."""
    tc = (entry.tcomment or '')
    return 'LEGAL DRAFT' in tc


def assign_block(entry):
    """Return (block_id, title, color, description) for a given entry."""
    if is_legal(entry):
        return BLOCKS[0]  # B7_LEGAL
    keys = reference_keys(entry)
    for block in BLOCKS[1:]:
        prefixes = block[3]
        if not prefixes:
            continue
        for k in keys:
            for p in prefixes:
                if p.endswith('_') and k.startswith(p):
                    return block
                if k == p or k.startswith(p):
                    # Avoid s_avanzosc_cta_contacto matching s_avanzosc_cta_kit_consulting
                    if k.startswith(p):
                        # Check exact or proper boundary
                        if len(k) == len(p) or k[len(p)] in ('_', ''):
                            return block
                        # also accept exact equality
                        if k == p:
                            return block
                        # otherwise continue (no match)
                        continue
    return FALLBACK_BLOCK


def context_text(entry):
    """Human-readable context string from references + comment."""
    keys = reference_keys(entry)
    # Map technical key to human label
    label_map = {
        # menus
        'menu_inicio': 'Menú: Inicio',
        'menu_industrial': 'Menú: Industrial',
        'menu_distribucion': 'Menú: Distribución',
        'menu_servicios': 'Menú: Servicios',
        'menu_academias': 'Menú: Academias',
        'menu_tienda': 'Menú: Tienda',
        'menu_formacion': 'Menú: Formación',
        'menu_conocenos': 'Menú: Conócenos',
        'menu_contacto': 'Menú: Contacto',
        # menu_trabaja eliminado post-v1 — entry retirada; mapping
        # se mantiene por si quedara alguna referencia residual.
        'menu_soluciones': 'Menú: Soluciones (dropdown)',
        # pages
        'page_industrial': '/industrial',
        'page_distribucion': '/distribucion',
        'page_servicios': '/servicios',
        'page_academias': '/academias',
        'page_conocenos': '/conocenos',
        # page_trabaja_con_nosotros eliminada post-v1.
        'page_contacto': '/contacto',
        'page_contacto_gracias': '/contacto/gracias',
        'page_aviso_legal': '/aviso-legal',
        'page_politica_privacidad': '/politica-privacidad',
        'page_politica_cookies': '/politica-cookies',
        'page_kit_consulting': '/kit-consulting',
        # snippets
        's_avanzosc_hero': 'Snippet: Hero (home + sectoriales)',
        's_avanzosc_pilares': 'Snippet: Pilares (home)',
        's_avanzosc_sectores': 'Snippet: Sectores grid (home)',
        's_avanzosc_contador': 'Snippet: Contador (home)',
        's_avanzosc_timeline': 'Snippet: Timeline (home)',
        's_avanzosc_equipo': 'Snippet: Equipo (home)',
        's_avanzosc_caso_exito': 'Snippet: Caso éxito (home + sectoriales)',
        's_avanzosc_sector_specifics': 'Snippet: Sector specifics (sectoriales)',
        's_avanzosc_cta_kit_consulting': 'Snippet: CTA Kit Consulting (home)',
        's_avanzosc_cta_contacto': 'Snippet: CTA Contacto (home + sectoriales + conócenos)',
        # header / footer
        'footer_avanzosc': 'Footer (todas las páginas)',
        'footer_copyright_avanzosc': 'Footer copyright (todas las páginas)',
        'header_acceso_clientes': 'Header: botón Acceso clientes',
        'header_mobile_buttons': 'Header: botones mobile (<992px)',
        'layout_mobile_overlay': 'Header: overlay mobile (<992px)',
    }
    labels = []
    seen = set()
    for k in keys:
        lbl = label_map.get(k, k)
        if lbl not in seen:
            labels.append(lbl)
            seen.add(lbl)
    ctx = ' · '.join(labels)
    # If translator comment has DRAFT description, append it
    tc = (entry.tcomment or '').strip()
    if tc:
        # Strip DRAFT markers, keep description after the dash
        for marker in ('DRAFT - REVIEW NEEDED', 'LEGAL DRAFT - REVIEW NEEDED'):
            if marker in tc:
                # Often format: "DRAFT - REVIEW NEEDED — <description>"
                parts = tc.split('—', 1)
                if len(parts) > 1:
                    note = parts[1].strip()
                    if note and note not in ctx:
                        ctx = f'{ctx} — {note}' if ctx else note
                break
    return ctx or '(sin contexto)'


def main():
    po = polib.pofile(PO_PATH)
    # Filter: entries marked fuzzy (DRAFT)
    fuzzy_entries = [e for e in po if 'fuzzy' in (e.flags or [])]
    # Also collect non-fuzzy entries that may have DRAFT markers (defensive — sub-bloque B
    # observation: some new strings might lack uniform marker). Match the canonical markers
    # used by the .po: '# DRAFT - REVIEW NEEDED' or '# LEGAL DRAFT - REVIEW NEEDED'.
    # Substring 'DRAFT' alone yields false positives on validated entries that mention
    # 'NO DRAFT' or 'pre-DRAFT' in their tcomment.
    def is_pending_marker(tc):
        return ('DRAFT - REVIEW NEEDED' in tc) or ('LEGAL DRAFT - REVIEW NEEDED' in tc)
    non_fuzzy_with_marker = [
        e for e in po
        if 'fuzzy' not in (e.flags or [])
        and is_pending_marker(e.tcomment or '')
    ]
    # Combine
    all_pending = fuzzy_entries + non_fuzzy_with_marker

    # Group by block
    by_block = {}
    for entry in all_pending:
        block = assign_block(entry)
        by_block.setdefault(block, []).append(entry)

    # Build workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ---- Sheet 0: INSTRUCCIONES ----
    ws0 = wb.create_sheet('INSTRUCCIONES', 0)
    write_instructions(ws0, by_block, len(all_pending), len(non_fuzzy_with_marker))

    # ---- Block sheets ----
    block_order = [b for b in BLOCKS] + [FALLBACK_BLOCK]
    for block in block_order:
        entries = by_block.get(block, [])
        if not entries:
            continue
        write_block_sheet(wb, block, entries)

    wb.save(OUT_PATH)
    print(f'OK · {len(all_pending)} pending strings · {len([b for b in by_block if by_block[b]])} blocks · {OUT_PATH}')

    # Counts by block (sorted by display order)
    print('\nConteo por bloque:')
    for block in block_order:
        entries = by_block.get(block, [])
        if entries:
            print(f'  {block[0]:20} {block[1]:42} {len(entries):4}')

    # Defensive: print non-fuzzy with marker (anomalies)
    if non_fuzzy_with_marker:
        print(f'\n[anomaly] {len(non_fuzzy_with_marker)} strings con DRAFT comment pero SIN flag fuzzy:')
        for e in non_fuzzy_with_marker[:5]:
            print(f'  {make_id(e.msgid)} {e.msgid[:60]!r}')


def write_instructions(ws, by_block, total_pending, anomaly_count):
    bold = Font(bold=True, size=11)
    title_font = Font(bold=True, size=14)
    h2_font = Font(bold=True, size=12, color='1B2A41')
    fill_yellow = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fill_red = PatternFill(start_color='FFB3B3', end_color='FFB3B3', fill_type='solid')

    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 100

    rows = [
        ('', 'Avanzosc · Q1 — Validación lingüística EU', title_font, None),
        ('', '', None, None),
        ('', f'Generado: 2026-04-30 · Total strings pendientes: {total_pending}', None, None),
        ('', '', None, None),
        ('', '1. ¿Qué es esto?', h2_font, None),
        ('', 'Las traducciones al euskera del nuevo sitio web de Avanzosc están actualmente en estado DRAFT (borrador automático generado por el dev). Necesitamos que el equipo de Avanzosc las revise y corrija lo que haga falta antes de publicar la web en producción.', None, None),
        ('', '', None, None),
        ('', '2. ¿Qué tienes que hacer?', h2_font, None),
        ('', 'Por cada fila de las pestañas siguientes:', None, None),
        ('', '  · Lee la columna ES (source) — el texto original en castellano.', None, None),
        ('', '  · Lee la columna EU DRAFT (current) — la traducción actual generada.', None, None),
        ('', '  · Si la traducción está bien: dejar la columna EU FINAL VACÍA. (Vacío = aprobamos la draft.)', None, None),
        ('', '  · Si la traducción debe corregirse: escribir la versión correcta en la columna EU FINAL.', None, None),
        ('', '  · Si tienes dudas, variantes alternativas o quieres comentar algo: usar la columna Notas.', None, None),
        ('', '  · La columna Contexto te dice DÓNDE aparece la string (URL/snippet) — útil para desambiguar duplicados.', None, None),
        ('', '', None, None),
        ('', '3. Glosario (NO traducir, mantener literal en EU también)', h2_font, None),
        ('', '  · Avanzosc · Odoo · OdooMRP · OCA · OpenERP · OpenUpgrade · TinyERP', None, None),
        ('', '  · Kit Consulting · Kit Digital (programas Red.es; nombres oficiales)', None, None),
        ('', '  · ERP · MRP · CRM · POS · EDI · ADR · AECOC · SILICIE · SAT (siglas técnicas)', None, None),
        ('', '  · CIF · S.L. (datos legales)', None, None),
        ('', '  · Direcciones de email, números de teléfono, URLs, IDs (todos literales)', None, None),
        ('', '', None, None),
        ('', '4. Tono y registro', h2_font, None),
        ('', 'B2B profesional cercano. Tutear (zu) coherente con el ES original. Evitar palabras coloquiales pero también evitar registro excesivamente formal/burocrático. Imitar el tono de la columna ES.', None, None),
        ('', '', None, None),
        ('', '5. Bloque Legal — IMPORTANTE', h2_font, None),
        ('', '⚠ La pestaña «Legal (Q3 — ESPERAR asesoría)» NO se revisa todavía. Esos textos son las páginas legales (aviso, privacidad, cookies) y dependen de revisión por la asesoría legal de Avanzosc (gate Q3). Cuando la asesoría valide los textos ES base, ese bloque se reabrirá. Por ahora dejarlo COMO ESTÁ (no editar).', fill_red, None),
        ('', '', None, None),
        ('', '6. Cómo devolver el archivo', h2_font, None),
        ('', '  · Guardar el XLSX con los cambios (mismo formato XLSX, NO PDF, NO Word).', None, None),
        ('', '  · Devolverlo al desarrollador (Aner / equipo dev). El dev parsea la columna EU FINAL, mergea al fichero de traducciones del módulo (i18n/eu.po), levanta el flag DRAFT y verifica que la web sigue funcionando.', None, None),
        ('', '  · El dev puede preguntar dudas concretas si una entrada en Notas necesita aclaración.', None, None),
        ('', '', None, None),
        ('', '7. Atención a duplicados', h2_font, None),
        ('', 'Si una misma string ES aparece en varias filas (puede ocurrir si está en varios sitios de la web), los IDs Q1-XXXXXX serán DISTINTOS pero el ES idéntico. Decisión: traducir igual en ambas filas. Si la columna Contexto sugiere matices, traducir distinto y anotar el porqué en Notas.', None, None),
        ('', '', None, None),
        ('', '8. Resumen visual', h2_font, None),
        ('', f'  · Strings pendientes total: {total_pending}', None, None),
    ]
    # Add per-block counts
    for block in [b for b in BLOCKS] + [FALLBACK_BLOCK]:
        entries = by_block.get(block, [])
        if entries:
            tag = '⚠ ESPERAR Q3' if block[0] == 'B7_LEGAL' else ''
            rows.append(('', f'  · {block[1]}: {len(entries)} strings  {tag}', None, fill_red if 'B7_LEGAL' in block[0] else None))

    if anomaly_count:
        rows.append(('', '', None, None))
        rows.append(('', f'(*) {anomaly_count} strings tienen comentario DRAFT pero sin flag fuzzy — incluidas en la pestaña que toque por contexto.', None, None))

    rows.append(('', '', None, None))
    rows.append(('', '9. Convención del color amarillo claro', h2_font, None))
    rows.append(('', 'Las celdas con fondo amarillo claro son las que tú rellenas (EU FINAL y Notas).', fill_yellow, None))

    instructions_col_widths = (4, 100)  # cols A and B widths in chars
    for i, (a, b, font, fill) in enumerate(rows, start=1):
        ws.cell(row=i, column=2, value=b)
        if font:
            ws.cell(row=i, column=2).font = font
        if fill:
            ws.cell(row=i, column=2).fill = fill
        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        # Calculate row height to fit wrapped text (only column B carries content here).
        ws.row_dimensions[i].height = compute_row_height(('', b), instructions_col_widths)


def write_block_sheet(wb, block, entries):
    block_id, title, color, _, description = block
    safe_name = title[:31]  # Excel sheet name limit
    # Strip illegal chars
    for c in '/\\?*[]:':
        safe_name = safe_name.replace(c, '-')
    ws = wb.create_sheet(safe_name)
    ws.sheet_properties.tabColor = color

    bold = Font(bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='1B2A41', end_color='1B2A41', fill_type='solid')
    fill_eu_final = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fill_notas = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    fill_legal = PatternFill(start_color='FFE5E5', end_color='FFE5E5', fill_type='solid')
    thin = Side(border_style='thin', color='CCCCCC')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # Title row
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14, color='1B2A41')
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    # Description row
    ws.cell(row=2, column=1, value=description).font = Font(italic=True, size=10, color='555555')
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    # Special legal warning
    if block_id == 'B7_LEGAL':
        ws.cell(row=3, column=1, value='⚠ ESPERAR Q3 ASESORÍA — bloque informativo, NO editar todavía').font = Font(bold=True, size=11, color='AA0000')
        ws.cell(row=3, column=1).fill = fill_legal
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        header_row = 4
    else:
        header_row = 3

    # Header row
    headers = ['ID', 'ES (source)', 'EU DRAFT (current)', 'EU FINAL', 'Notas', 'Contexto']
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=header_row, column=col, value=h)
        c.font = bold
        c.fill = header_fill
        c.alignment = Alignment(wrap_text=True, vertical='center', horizontal='left')

    # Column widths
    widths = [12, 50, 50, 50, 30, 45]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Set heights for the title/description/banner rows that sit before the
    # column header. Their content can be long (e.g. the LEGAL banner is one
    # full sentence). Compute against width of column A merged across 6 cols
    # (sum of widths = 12+50+50+50+30+45 = 237 chars).
    merged_width_chars = sum(widths)
    ws.row_dimensions[1].height = compute_row_height((title,) + ('',) * 5,
                                                      (merged_width_chars,) + (0,) * 5)
    ws.row_dimensions[2].height = compute_row_height((description,) + ('',) * 5,
                                                      (merged_width_chars,) + (0,) * 5)
    if block_id == 'B7_LEGAL':
        ws.row_dimensions[3].height = compute_row_height(
            ('⚠ ESPERAR Q3 ASESORÍA — bloque informativo, NO editar todavía',) + ('',) * 5,
            (merged_width_chars,) + (0,) * 5)
    # Header row needs default-ish height (column titles are short).
    ws.row_dimensions[header_row].height = LINE_HEIGHT_PT + PADDING_PT

    # Sort entries by msgid for stable order
    entries_sorted = sorted(entries, key=lambda e: e.msgid.lower())

    # Data rows
    for row_idx, entry in enumerate(entries_sorted, start=header_row + 1):
        cell_values = (
            make_id(entry.msgid),
            entry.msgid,
            entry.msgstr,
            '',
            '',
            context_text(entry),
        )
        for col, v in enumerate(cell_values, start=1):
            ws.cell(row=row_idx, column=col, value=v)

        # Yellow fill on EU FINAL + Notas
        ws.cell(row=row_idx, column=4).fill = fill_eu_final
        ws.cell(row=row_idx, column=5).fill = fill_notas

        # Wrap text + top alignment + thin borders
        for col in range(1, 7):
            cell = ws.cell(row=row_idx, column=col)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.border = border

        # Calculate row height: max wrapped-line count across the 6 columns.
        ws.row_dimensions[row_idx].height = compute_row_height(cell_values, tuple(widths))

    # Freeze header row
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)


if __name__ == '__main__':
    main()
